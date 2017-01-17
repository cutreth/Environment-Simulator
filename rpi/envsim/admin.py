from django.contrib import admin
from django.http import HttpResponse

from models import Reading
from models import Config


def export(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Readings.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Readings"

    row_num = 0

    columns = [
        (u"Instant", 20),
        (u"Temp Val", 10),
        (u"Humid Val", 10),
        (u"Temp State", 12),
        (u"Humid State", 12),
        (u"Light State", 12),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.instant,
            obj.temp_val,
            obj.humid_val,
            obj.temp_state,
            obj.humid_state,
            obj.light_state,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response

export.short_description = u"Export selected readings"


class ReadingAdmin(admin.ModelAdmin):

    fieldsets = [
        ('Values', {'fields': ['instant', 'temp_val', 'humid_val']}),
        ('States', {'fields': ['temp_state', 'humid_state', 'light_state']}),
    ]

    readonly_fields = ('instant', 'temp_val', 'humid_val', 'temp_state', 'humid_state', 'light_state')

    date_hierarchy = 'instant'

    actions = [export]


class ConfigAdmin(admin.ModelAdmin):

    fieldsets = [
        ('Values', {'fields': ['temp_low', 'temp_high', 'humid_low', 'humid_high', 'hour_morning', 'hour_night']}),
        ('States', {'fields': ['temp_state', 'humid_state', 'light_state']}),
        ('Pins', {'fields': ['temp_humid_sensor', 'temp_pin', 'humid_pin', 'light_pin']}),
    ]

    readonly_fields = ('instant', )

admin.site.register(Reading, ReadingAdmin)
admin.site.register(Config, ConfigAdmin)
